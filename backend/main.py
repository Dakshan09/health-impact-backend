from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from tools.prediction.health_impact_predictor import generate_health_prediction, PredictionRequest
from tools.report_generation.medical_report_generator import generate_medical_report
from tools.report_generation.pdf_compiler import compile_pdf, generate_output_path
from tools.visualization.health_dashboard_creator import create_risk_visualization
from tools.visualization.visualization_builder import build_visualization_html
from tools.intervention.schedule_generator import generate_intervention_schedule
import uvicorn
import os
import io
import json
import tempfile
import traceback
import sys
sys.stdout.reconfigure(encoding='utf-8')
from dotenv import load_dotenv
from pydantic import BaseModel
from typing import Optional
from datetime import datetime

# Force load .env from the backend directory
env_path = os.path.join(os.path.dirname(__file__), ".env")
load_dotenv(env_path)

app = FastAPI(title="Health Impact Analyzer API")

# ─── CORS ─────────────────────────────────────────────────────────────────────
# Allow all Vercel preview deployments (*.vercel.app) plus local dev
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest
from starlette.responses import Response as StarletteResponse
import re

ALLOWED_ORIGIN_PATTERNS = [
    r"^http://localhost(:\d+)?$",
    r"^http://127\.0\.0\.1(:\d+)?$",
    r"^https://[\w-]+(\.[\w-]+)*\.vercel\.app$",
    r"^https://antigravity-health\.vercel\.app$",
    r"^https://health-impact-analyzer\.vercel\.app$",
]

def _is_allowed_origin(origin: str) -> bool:
    return any(re.match(p, origin) for p in ALLOWED_ORIGIN_PATTERNS)

class DynamicCORSMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: StarletteRequest, call_next):
        origin = request.headers.get("origin", "")
        allowed = _is_allowed_origin(origin)

        if request.method == "OPTIONS":
            response = StarletteResponse(status_code=204)
        else:
            response = await call_next(request)

        if allowed:
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "*"
            response.headers["Access-Control-Expose-Headers"] = "content-disposition"
        return response

app.add_middleware(DynamicCORSMiddleware)

# ─── Predict ────────────────────────────────────────────────────────────────

def _sanitize_patient_data(patient_data: dict) -> dict:
    """
    Normalise patient_data from the frontend before passing to PredictionRequest.
    - Converts list fields to comma-separated strings.
    - Maps frontend field names to model field names.
    - Removes None values to prevent validation errors.
    """
    d = dict(patient_data)  # shallow copy

    # Flatten list → string fields
    for key in ("familyHistory", "currentSymptoms", "medications", "allergies", "surgeries", "concerns"):
        val = d.get(key)
        if isinstance(val, list):
            d[key] = ", ".join(str(x) for x in val)
        elif val is None:
            d[key] = ""

    # Frontend sends existingConditions, model expects conditions
    if "existingConditions" in d and "conditions" not in d:
        val = d["existingConditions"]
        d["conditions"] = val if isinstance(val, list) else [val] if val else []

    # Frontend sends dietType, model expects dietQuality
    if "dietType" in d and "dietQuality" not in d:
        d["dietQuality"] = d["dietType"]

    # Filter to only fields the model knows about, dropping Nones and coercing to appropriate types
    sanitized = {}
    for k, v in d.items():
        if k in PredictionRequest.model_fields:
            if v is None:
                continue # Skip None to use Pydantic defaults
            if k == "conditions":
                sanitized[k] = v if isinstance(v, list) else [v] if v else []
            else:
                sanitized[k] = str(v)

    return sanitized


@app.post("/api/predict")
async def predict_health(request: PredictionRequest):
    data = request.model_dump()
    analysis = generate_health_prediction(data)
    return {"analysis": analysis}

@app.get("/health")
def health_check():
    return {"status": "ok"}

@app.get("/api/test-smtp")
def test_smtp():
    """Diagnostic endpoint to test SMTP configuration."""
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_email = os.getenv("SMTP_EMAIL", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    result = {
        "smtp_host": smtp_host,
        "smtp_port": smtp_port,
        "smtp_email": smtp_email,
        "smtp_password_set": bool(smtp_password),
        "smtp_password_length": len(smtp_password),
        "methods_tried": [],
    }
    if not smtp_email or not smtp_password:
        result["error"] = "SMTP_EMAIL or SMTP_PASSWORD not configured"
        return result
    import smtplib
    # Try STARTTLS on port 587
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(smtp_email, smtp_password)
            result["methods_tried"].append({"method": "starttls", "port": smtp_port, "status": "OK"})
            result["authenticated"] = True
            result["status"] = "OK - STARTTLS login successful"
            return result
    except Exception as e:
        result["methods_tried"].append({"method": "starttls", "port": smtp_port, "error": str(e)})
    # Try SSL on port 465
    try:
        with smtplib.SMTP_SSL(smtp_host, 465, timeout=15) as server:
            server.login(smtp_email, smtp_password)
            result["methods_tried"].append({"method": "ssl", "port": 465, "status": "OK"})
            result["authenticated"] = True
            result["status"] = "OK - SSL login successful"
            return result
    except Exception as e:
        result["methods_tried"].append({"method": "ssl", "port": 465, "error": str(e)})
    result["status"] = "FAILED - All methods failed"
    result["authenticated"] = False
    return result

# ─── Report models ───────────────────────────────────────────────────────────

class ReportRequest(BaseModel):
    patientData: dict
    analysis: Optional[str] = None

# ─── Submit (full pipeline) ───────────────────────────────────────────────────

@app.post("/api/submit")
async def submit_assessment(request: ReportRequest):
    """
    Full pipeline:
    1. Run AI prediction
    2. Generate all 3 reports (Clinical PDF, Visualization, Schedule Excel)
    3. Save assessment + reports to Supabase
    4. Send email with all reports as attachments
    5. Return summary JSON
    """
    patient_data = request.patientData
    analysis = request.analysis

    patient_name = patient_data.get("fullName") or "Patient"
    patient_email = patient_data.get("email") or ""
    safe_name = "".join(c if c.isalnum() else "_" for c in patient_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp_dir = os.path.join(tempfile.gettempdir(), "user_reports")
    os.makedirs(tmp_dir, exist_ok=True)

    # ── Step 1: AI Prediction ─────────────────────────────────────────────
    if not analysis:
        try:
            pred_request = PredictionRequest(**_sanitize_patient_data(patient_data))
            analysis = generate_health_prediction(pred_request.model_dump())
        except Exception as e:
            analysis = f"AI analysis unavailable: {str(e)}"

    # ── Step 2: Risk Visualization Data ───────────────────────────────────
    dashboard = create_risk_visualization(analysis, patient_data)
    risk_categories = dashboard.get("risk_categories", {})
    lifestyle_scores = dashboard.get("lifestyle_scores", {})
    overall_risk = dashboard.get("overall_risk", "Unknown")

    # ── Step 3: Generate Clinical PDF ────────────────────────────────────
    clinical_path = ""
    try:
        report_md = generate_medical_report(
            prediction_text=analysis,
            patient_name=patient_name,
            patient_data=patient_data,
            risk_categories=risk_categories,
        )
        clinical_output = os.path.join(tmp_dir, f"clinical_{safe_name}_{timestamp}.pdf")
        clinical_path = compile_pdf(report_md, clinical_output)
    except Exception as e:
        print(f"WARN: Clinical report generation failed: {e}".encode("ascii", "replace").decode("ascii"))
        traceback.print_exc()

    # ── Step 4: Generate Visualization Report (HTML) ─────────────────────
    visualization_path = ""
    try:
        html_content = build_visualization_html(
            patient_name=patient_name,
            patient_data=patient_data,
            risk_categories=risk_categories,
            lifestyle_scores=lifestyle_scores,
            overall_risk=overall_risk,
            analysis=analysis,
        )
        html_path = os.path.join(tmp_dir, f"visualization_{safe_name}_{timestamp}.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        visualization_path = html_path  # Always HTML
    except Exception as e:
        print(f"WARN: Visualization report generation failed: {e}".encode("ascii", "replace").decode("ascii"))
        traceback.print_exc()

    # ── Step 5: Generate Schedule Excel/PDF ──────────────────────────────
    schedule_path = ""
    try:
        schedule = generate_intervention_schedule(
            patient_data=patient_data,
            risk_categories=risk_categories,
            duration_days=30,
        )
        xlsx_path = os.path.join(tmp_dir, f"schedule_{safe_name}_{timestamp}.xlsx")
        _build_schedule_excel(schedule, xlsx_path)
        schedule_path = xlsx_path
    except ImportError:
        # fallback to PDF
        try:
            schedule_md = _build_schedule_markdown(schedule)
            pdf_path = os.path.join(tmp_dir, f"schedule_{safe_name}_{timestamp}.pdf")
            schedule_path = compile_pdf(schedule_md, pdf_path)
        except Exception as e2:
            print(f"WARN: Schedule report generation failed: {e2}".encode("ascii", "replace").decode("ascii"))
            traceback.print_exc()
    except Exception as e:
        print(f"WARN: Schedule report generation failed: {e}".encode("ascii", "replace").decode("ascii"))
        traceback.print_exc()

    # ── Step 6: Save to Supabase (non-blocking — never stops email) ──────
    assessment_id = ""
    try:
        from tools.data_management.supabase_client import save_assessment
        assessment_id = save_assessment(patient_data, analysis, overall_risk)
    except Exception as e:
        print(f"WARN: Supabase save_assessment failed (continuing): {str(e)[:200]}")

    # ── Step 7: Send Email with Attachments ──────────────────────────────
    # This ALWAYS runs regardless of Supabase success/failure above
    email_sent = False
    if patient_email:
        try:
            from tools.communication.email_delivery import send_reports_email
            attachments = [p for p in [clinical_path, visualization_path, schedule_path] if p and os.path.exists(p)]
            print(f"INFO: Sending email to {patient_email} with {len(attachments)} attachment(s)")
            print(f"INFO: Attachment paths: {attachments}")
            email_sent = send_reports_email(
                to_address=patient_email,
                patient_name=patient_name,
                analysis_summary=analysis,
                overall_risk=overall_risk,
                attachments=attachments,
            )
            print(f"INFO: email_sent = {email_sent}")
        except Exception as e:
            print(f"WARN: Email sending failed: {str(e)[:300]}")
            traceback.print_exc()
    else:
        print("WARN: No patient email provided — skipping email send")

    # ── Step 8: Save Report Record to Supabase (non-blocking) ─────────────
    try:
        from tools.data_management.supabase_client import save_report_record
        save_report_record(
            assessment_id=assessment_id,
            patient_name=patient_name,
            patient_email=patient_email,
            clinical_path=clinical_path,
            schedule_path=schedule_path,
            visualization_path=visualization_path,
            email_sent=email_sent,
        )
    except Exception as e:
        print(f"WARN: Supabase save_report_record failed (non-blocking): {str(e)[:200]}")

    # ─── Return summary ───────────────────────────────────────────────────
    return {
        "success": True,
        "assessment_id": assessment_id,
        "overall_risk": overall_risk,
        "email_sent": email_sent,
        "email_address": patient_email,
        "analysis": analysis,
        "reports": {
            "clinical": os.path.basename(clinical_path) if clinical_path else None,
            "visualization": os.path.basename(visualization_path) if visualization_path else None,
            "schedule": os.path.basename(schedule_path) if schedule_path else None,
        },
        "risk_categories": risk_categories,
        "lifestyle_scores": lifestyle_scores,
    }

# ─── Email Reports (standalone — lighter than /api/submit) ────────────────────

class EmailReportRequest(BaseModel):
    patientData: dict
    analysis: Optional[str] = None
    email: Optional[str] = None  # override patient email

@app.post("/api/email-reports")
async def email_reports(request: EmailReportRequest):
    """Generate reports and email them — standalone endpoint for Dashboard use."""
    patient_data = request.patientData
    analysis = request.analysis or ""
    patient_name = patient_data.get("fullName") or "Patient"
    patient_email = request.email or patient_data.get("email") or ""

    if not patient_email:
        raise HTTPException(status_code=400, detail="No email address provided")

    safe_name = "".join(c if c.isalnum() else "_" for c in patient_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp_dir = os.path.join(tempfile.gettempdir(), "user_reports")
    os.makedirs(tmp_dir, exist_ok=True)

    # AI Prediction if not provided
    if not analysis:
        try:
            pred_request = PredictionRequest(**_sanitize_patient_data(patient_data))
            analysis = generate_health_prediction(pred_request.model_dump())
        except Exception as e:
            analysis = f"AI analysis unavailable: {str(e)}"

    # Risk data
    dashboard = create_risk_visualization(analysis, patient_data)
    risk_categories = dashboard.get("risk_categories", {})
    lifestyle_scores = dashboard.get("lifestyle_scores", {})
    overall_risk = dashboard.get("overall_risk", "Unknown")

    # Generate Clinical PDF
    clinical_path = ""
    try:
        report_md = generate_medical_report(
            prediction_text=analysis, patient_name=patient_name,
            patient_data=patient_data, risk_categories=risk_categories,
        )
        clinical_path = compile_pdf(report_md, os.path.join(tmp_dir, f"clinical_{safe_name}_{timestamp}.pdf"))
    except Exception as e:
        print(f"WARN: Clinical report failed: {e}")

    # Generate Visualization HTML
    visualization_path = ""
    try:
        html_content = build_visualization_html(
            patient_name=patient_name, patient_data=patient_data,
            risk_categories=risk_categories, lifestyle_scores=lifestyle_scores,
            overall_risk=overall_risk, analysis=analysis,
        )
        html_path = os.path.join(tmp_dir, f"visualization_{safe_name}_{timestamp}.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        visualization_path = html_path
    except Exception as e:
        print(f"WARN: Visualization report failed: {e}")

    # Generate Schedule Excel
    schedule_path = ""
    try:
        schedule = generate_intervention_schedule(
            patient_data=patient_data, risk_categories=risk_categories, duration_days=30,
        )
        xlsx_path = os.path.join(tmp_dir, f"schedule_{safe_name}_{timestamp}.xlsx")
        _build_schedule_excel(schedule, xlsx_path)
        schedule_path = xlsx_path
    except Exception as e:
        print(f"WARN: Schedule report failed: {e}")

    # Send email
    from tools.communication.email_delivery import send_reports_email
    attachments = [p for p in [clinical_path, visualization_path, schedule_path] if p and os.path.exists(p)]
    email_sent = send_reports_email(
        to_address=patient_email, patient_name=patient_name,
        analysis_summary=analysis, overall_risk=overall_risk, attachments=attachments,
    )

    return {
        "success": email_sent,
        "email_sent": email_sent,
        "email_address": patient_email,
        "attachments_count": len(attachments),
    }

# ─── Clinical PDF Report ─────────────────────────────────────────────────────

@app.post("/api/report/clinical")
async def generate_clinical_report(request: ReportRequest):
    """Generate a clinical PDF report from patient data + AI analysis."""
    patient_data = request.patientData
    analysis = request.analysis

    # Run prediction if not provided
    if not analysis:
        pred_request = PredictionRequest(**_sanitize_patient_data(patient_data))
        analysis = generate_health_prediction(pred_request.model_dump())

    # Build risk categories from patient data
    dashboard = create_risk_visualization(analysis, patient_data)
    risk_categories = dashboard.get("risk_categories", {})

    # Generate markdown report
    patient_name = patient_data.get("fullName") or "Patient"
    report_md = generate_medical_report(
        prediction_text=analysis,
        patient_name=patient_name,
        patient_data=patient_data,
        risk_categories=risk_categories,
    )

    # Compile to PDF
    tmp_dir = os.path.join(tempfile.gettempdir(), "user_reports")
    os.makedirs(tmp_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() else "_" for c in patient_name)
    output_path = os.path.join(tmp_dir, f"clinical_{safe_name}_{timestamp}.pdf")

    result_path = compile_pdf(report_md, output_path)

    ext = os.path.splitext(result_path)[1]  # .pdf or .html
    media_type = "application/pdf" if ext == ".pdf" else "text/html"
    return FileResponse(
        path=result_path,
        filename=f"Clinical_Report_{safe_name}{ext}",
        media_type=media_type,
    )

# ─── Visualization PDF Report ─────────────────────────────────────────────────

@app.post("/api/report/visualization")
async def generate_visualization_report(request: ReportRequest):
    """Generate a visual HTML/PDF report with risk charts."""
    patient_data = request.patientData
    analysis = request.analysis

    if not analysis:
        pred_request = PredictionRequest(**_sanitize_patient_data(patient_data))
        analysis = generate_health_prediction(pred_request.model_dump())

    # Build dashboard data
    dashboard = create_risk_visualization(analysis, patient_data)
    risk_categories = dashboard.get("risk_categories", {})
    lifestyle_scores = dashboard.get("lifestyle_scores", {})
    overall_risk = dashboard.get("overall_risk", "Unknown")
    patient_name = patient_data.get("fullName") or "Patient"

    # Generate rich HTML visualization report
    html_content = build_visualization_html(
        patient_name=patient_name,
        patient_data=patient_data,
        risk_categories=risk_categories,
        lifestyle_scores=lifestyle_scores,
        overall_risk=overall_risk,
        analysis=analysis,
    )

    # Always return as HTML
    tmp_dir = os.path.join(tempfile.gettempdir(), "user_reports")
    os.makedirs(tmp_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() else "_" for c in patient_name)
    html_path = os.path.join(tmp_dir, f"visualization_{safe_name}_{timestamp}.html")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    return FileResponse(
        path=html_path,
        filename=f"Visualization_Report_{safe_name}.html",
        media_type="text/html",
    )

# ─── Schedule Report (PDF + Excel) ───────────────────────────────────────────

@app.post("/api/report/schedule")
async def generate_schedule_report(request: ReportRequest):
    """Generate a 30-day intervention schedule as Excel or PDF."""
    patient_data = request.patientData
    analysis = request.analysis or ""

    dashboard = create_risk_visualization(analysis, patient_data)
    risk_categories = dashboard.get("risk_categories", {})

    schedule = generate_intervention_schedule(
        patient_data=patient_data,
        risk_categories=risk_categories,
        duration_days=30,
    )

    patient_name = patient_data.get("fullName") or "Patient"
    safe_name = "".join(c if c.isalnum() else "_" for c in patient_name)
    tmp_dir = os.path.join(tempfile.gettempdir(), "user_reports")
    os.makedirs(tmp_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Try Excel first
    try:
        import openpyxl
        xlsx_path = os.path.join(tmp_dir, f"schedule_{safe_name}_{timestamp}.xlsx")
        _build_schedule_excel(schedule, xlsx_path)
        return FileResponse(
            path=xlsx_path,
            filename=f"Intervention_Schedule_{safe_name}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ImportError:
        pass

    # Fallback: PDF schedule
    schedule_md = _build_schedule_markdown(schedule)
    pdf_path = os.path.join(tmp_dir, f"schedule_{safe_name}_{timestamp}.pdf")
    result_path = compile_pdf(schedule_md, pdf_path)
    return FileResponse(
        path=result_path,
        filename=f"Intervention_Schedule_{safe_name}.pdf",
        media_type="application/pdf",
    )


# ─── Schedule Excel Builder ───────────────────────────────────────────────────

def _build_schedule_excel(schedule: dict, output_path: str):
    """Build 30-day intervention schedule as Excel workbook with 6 sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Theme colours ─────────────────────────────────────────────────────
    TEAL_DARK = "0D7377"
    TEAL_MED  = "14919B"
    TEAL_LIGHT = "E0F7FA"
    ALT_ROW    = "F5F5F5"
    WHITE_FG   = "FFFFFF"

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header(ws, cell_ref, text, bg=TEAL_DARK, fg=WHITE_FG, size=11, bold=True):
        ws[cell_ref] = text
        ws[cell_ref].font = Font(bold=bold, color=fg, size=size)
        ws[cell_ref].fill = PatternFill("solid", fgColor=bg)
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_section(ws, cell_ref, text, bg=TEAL_MED, fg=WHITE_FG, size=11):
        ws[cell_ref] = text
        ws[cell_ref].font = Font(bold=True, color=fg, size=size)
        ws[cell_ref].fill = PatternFill("solid", fgColor=bg)
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")

    def style_category(ws, cell_ref, text, bg=TEAL_LIGHT):
        ws[cell_ref] = text
        ws[cell_ref].font = Font(bold=True, size=10, color="0D3B3E")
        ws[cell_ref].fill = PatternFill("solid", fgColor=bg)

    def style_cell(ws, cell_ref, text, bold=False, size=10, wrap=True):
        ws[cell_ref] = text
        ws[cell_ref].font = Font(bold=bold, size=size)
        ws[cell_ref].alignment = Alignment(wrap_text=wrap, vertical="top")

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: Overview
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 32

    # Title
    ws1.merge_cells("A1:D1")
    style_header(ws1, "A1", "30-Day Health Transformation", size=14)
    ws1.row_dimensions[1].height = 30

    # Dates
    row = 3
    style_cell(ws1, f"A{row}", "Start", bold=True)
    style_cell(ws1, f"B{row}", "Your Day 1")
    style_cell(ws1, f"C{row}", "End", bold=True)
    style_cell(ws1, f"D{row}", "Day 30")
    row += 1
    style_cell(ws1, f"A{row}", "Duration", bold=True)
    style_cell(ws1, f"B{row}", "30 days (4 weeks)")

    # Core Principles
    row = 6
    ws1.merge_cells(f"A{row}:D{row}")
    style_section(ws1, f"A{row}", "Core Principles")
    row += 1
    for principle in schedule.get("core_principles", []):
        style_cell(ws1, f"A{row}", f"  \u2713  {principle}")
        row += 1

    # Milestones table
    row += 1
    ws1.merge_cells(f"A{row}:D{row}")
    style_section(ws1, f"A{row}", "Milestones")
    row += 1
    for col, label in [("A", "Day"), ("B", "Checkpoint"), ("C", "Actions")]:
        style_header(ws1, f"{col}{row}", label, bg=TEAL_MED, size=10)
    row += 1
    for ms in schedule.get("milestones", []):
        style_cell(ws1, f"A{row}", str(ms["day"]))
        style_cell(ws1, f"B{row}", ms["checkpoint"])
        style_cell(ws1, f"C{row}", ms["actions"])
        row += 1

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: Daily Schedule
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Daily Schedule")
    daily_headers = [
        "Day", "Week", "Phase", "Breakfast", "Lunch", "Dinner",
        "Snacks", "Hydration Target", "Daily Focus",
        "Vegetables \u2713", "Fruit \u2713", "Water \u2713", "Activity \u2713", "Mindful Eating \u2713",
    ]
    col_widths = [6, 6, 20, 32, 32, 32, 28, 18, 32, 12, 10, 10, 10, 14]
    for idx, (header, width) in enumerate(zip(daily_headers, col_widths), start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        ws2.column_dimensions[col_letter].width = width
        style_header(ws2, f"{col_letter}1", header, size=9)

    # Data rows
    for row_idx, day in enumerate(schedule.get("daily_schedule", []), start=2):
        vals = [
            str(day["day"]), str(day["week"]), day["phase"],
            day["breakfast"], day["lunch"], day["dinner"],
            day["snacks"], day["hydration"], day["daily_focus"],
            "\u2610", "\u2610", "\u2610", "\u2610", "\u2610",
        ]
        for col_idx, val in enumerate(vals, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws2[f"{col_letter}{row_idx}"] = val
            ws2[f"{col_letter}{row_idx}"].font = Font(size=9)
            ws2[f"{col_letter}{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws2[f"{col_letter}{row_idx}"].border = thin_border
            # Alternating row colours
            if row_idx % 2 == 0:
                ws2[f"{col_letter}{row_idx}"].fill = PatternFill("solid", fgColor=ALT_ROW)

    # ══════════════════════════════════════════════════════════════════════
    # Sheets 3-6: Week 1 through Week 4
    # ══════════════════════════════════════════════════════════════════════
    for week_data in schedule.get("weeks", []):
        ws = wb.create_sheet(f"Week {week_data['number']}")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 28

        row = 1
        # Week title
        ws.merge_cells(f"A{row}:C{row}")
        style_header(ws, f"A{row}", week_data["title"], size=13)
        ws.row_dimensions[row].height = 28
        row += 2

        # ── Weekly Goals ──────────────────────────────────────────────
        ws.merge_cells(f"A{row}:C{row}")
        style_section(ws, f"A{row}", "Weekly Goals")
        row += 1
        for goal in week_data["goals"]:
            style_cell(ws, f"A{row}", "\u2610")
            style_cell(ws, f"B{row}", goal)
            row += 1
        row += 1

        # ── Shopping List ─────────────────────────────────────────────
        ws.merge_cells(f"A{row}:C{row}")
        style_section(ws, f"A{row}", "Shopping List")
        row += 1
        for category, items in week_data.get("shopping_list", {}).items():
            style_category(ws, f"A{row}", category)
            row += 1
            for item in items:
                style_cell(ws, f"A{row}", "  \u2022", size=9)
                style_cell(ws, f"B{row}", item, size=9)
                row += 1
        row += 1

        # ── Meal Prep ─────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:C{row}")
        style_section(ws, f"A{row}", "Meal Prep (Sunday or your preferred prep day)")
        row += 1
        for task in week_data.get("meal_prep", []):
            style_cell(ws, f"A{row}", "\u2610")
            style_cell(ws, f"B{row}", task)
            row += 1
        row += 1

        # ── Weekly Check-In ───────────────────────────────────────────
        ws.merge_cells(f"A{row}:C{row}")
        style_section(ws, f"A{row}", "Weekly Check-In")
        row += 1
        for question in week_data.get("check_in_questions", []):
            style_cell(ws, f"A{row}", "?")
            style_cell(ws, f"B{row}", question)
            row += 1

    wb.save(output_path)
    print(f"[OK] Schedule Excel saved to: {output_path}")


# ─── Schedule Markdown Builder ────────────────────────────────────────────────

def _build_schedule_markdown(schedule: dict) -> str:
    md = f"""# 30-Day Health Transformation
**Patient:** {schedule['patient_name']}
**Period:** {schedule['start_date']} \u2192 {schedule['end_date']}

---

## Core Principles

"""
    for p in schedule.get("core_principles", []):
        md += f"- \u2713 {p}\n"

    md += "\n---\n\n## Milestones\n\n"
    for ms in schedule.get("milestones", []):
        md += f"- **Day {ms['day']}**: {ms['checkpoint']} \u2014 {ms['actions']}\n"

    md += "\n---\n\n## Weekly Overview\n\n"
    for wk in schedule.get("weeks", []):
        md += f"### {wk['title']}\n\n"
        md += "**Goals:**\n"
        for g in wk["goals"]:
            md += f"- {g}\n"
        md += "\n"

    md += f"\n---\n\n*Generated by Health Impact Analyzer on {datetime.now().strftime('%Y-%m-%d')}*\n"
    return md


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
